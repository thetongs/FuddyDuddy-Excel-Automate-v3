## Excel Automation
# Load Libraries
import pandas as pd
import openpyxl
from openpyxl.styles.alignment import Alignment
import os

# 
def excel_door_open(s_filepath, d_filepath):
    # Load Datafile
    data = pd.read_excel(s_filepath, 
        sheet_name = "Door Matrix", 
        usecols = [1])

    # Get Door Numbers
    data = data[3:]
    data.columns = ["Door Numbers"]

    # Load Source Worksheet
    book = openpyxl.load_workbook(s_filepath)
    source = book["Checklist"]

    # For each door number
    for number in data["Door Numbers"]:
        # Create Worksheet
        book.copy_worksheet(source)
        book["Checklist Copy"].title = str(number)     
        book["{}".format(str(number))].print_area = "A1:N22"   
        
    book.save(d_filepath + "/" +"Result.xlsx")

    book = openpyxl.load_workbook(d_filepath + "/" + "Result.xlsx")
    n = book["Door Matrix"].max_row
    
    for r, door_n in zip(range(5, book["Door Matrix"].max_row + 1) , data["Door Numbers"]):
        book["Door Matrix"].cell(row = r, column = 2).hyperlink = ("#{}!A1".format(door_n))
        book["Door Matrix"].cell(row = r, column = 2).style = 'Hyperlink'
        book["Door Matrix"].cell(row = r, column = 2).alignment = Alignment(horizontal="center")
        
    book.save(d_filepath + "/" + "Result_hyperlinked.xlsx")

    book = openpyxl.load_workbook(d_filepath + "/"+ "Result_hyperlinked.xlsx")

    for sheet_name in book.sheetnames[2:]:
        # Add Hyperlink to 'Door Matrix'
        
        book[sheet_name]["O6"].value = '=HYPERLINK("#\'Door Matrix\'!A1","HOME")'
        book[sheet_name]["O6"].style = 'Hyperlink'
        book[sheet_name]["O6"].alignment = Alignment(horizontal="center")

        #
        book[sheet_name]["F3"].value = sheet_name

        # Add Formula to B5
        book[sheet_name]['B3'].value = "='Door Matrix'!B2"
        book[sheet_name]['B4'].value = "=VLOOKUP(F3,'Door Matrix'!B5:E{},3,0)".format(n)
        book[sheet_name]['B5'].value = "=VLOOKUP(F3,'Door Matrix'!B5:E{},4,0)".format(n)
        book[sheet_name]['F4'].value = "=VLOOKUP(F3,'Door Matrix'!B5:E{},2,0)".format(n)

    # Save Result
    file_name = os.path.basename(s_filepath).split('.')[0]

    book.save(d_filepath + "/" + "{}_Updated.xlsx".format(file_name))
    book.close()

    # Remove Temp Files
    os.remove(d_filepath + "/" + "Result.xlsx")
    os.remove(d_filepath + "/" + "Result_hyperlinked.xlsx")